import json, re, pathlib, shutil, unicodedata, time, math
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook
import orjson

base_dir = pathlib.Path('/mnt/data/dashboard_monitoramento_amazul_v2')
if base_dir.exists():
    shutil.rmtree(base_dir)
for sub in ['css','js','data/processed','data/geo','metodologia','scripts']:
    (base_dir/sub).mkdir(parents=True, exist_ok=True)

fne_path = pathlib.Path('/mnt/data/FNE_AnexoI_Contratações_032026.xlsx')
tipo_path = pathlib.Path('/mnt/data/Tipologia Amazônia Azul (v5).xlsx')

macro_map = {
    'AC':'Norte','AP':'Norte','AM':'Norte','PA':'Norte','RO':'Norte','RR':'Norte','TO':'Norte',
    'AL':'Nordeste','BA':'Nordeste','CE':'Nordeste','MA':'Nordeste','PB':'Nordeste','PE':'Nordeste','PI':'Nordeste','RN':'Nordeste','SE':'Nordeste',
    'DF':'Centro-Oeste','GO':'Centro-Oeste','MT':'Centro-Oeste','MS':'Centro-Oeste',
    'ES':'Sudeste','MG':'Sudeste','RJ':'Sudeste','SP':'Sudeste',
    'PR':'Sul','RS':'Sul','SC':'Sul'
}

def is_missing(x):
    if x is None: return True
    if isinstance(x, float) and math.isnan(x): return True
    return False

def clean_code(x):
    if is_missing(x): return ''
    s = str(x).strip(); s = re.sub(r'\.0$', '', s); s = re.sub(r'\D', '', s)
    return s.zfill(7) if s else ''

def norm_text(x, default='Não informado'):
    if is_missing(x): return default
    s = str(x).strip()
    if s == '' or s.lower() in ['nan', 'none', 'null', '-', 'na']:
        return default
    return s

def title_text(x):
    s = norm_text(x)
    return s.title() if s != 'Não informado' else s

def yes_no(x):
    s = norm_text(x, 'Não')
    sn = unicodedata.normalize('NFKD', s).encode('ASCII','ignore').decode('ASCII').lower().strip()
    if sn in ['sim','s','yes','1','true']: return 'Sim'
    if sn in ['nao','n','no','0','false']: return 'Não'
    return s

def sex_std(x):
    s = norm_text(x)
    sn = unicodedata.normalize('NFKD', s).encode('ASCII','ignore').decode('ASCII').upper().strip()
    if sn in ['F','FEMININO','MULHER','MULHERES']: return 'Mulheres'
    if sn in ['M','MASCULINO','HOMEM','HOMENS']: return 'Homens'
    return 'Não informado / pessoa jurídica / não aplicável'

def to_float(x):
    if is_missing(x): return 0.0
    if isinstance(x, (int, float)): return float(x)
    s = str(x).strip().replace('.', '').replace(',', '.') if ',' in str(x) else str(x).strip()
    try: return float(s)
    except Exception: return 0.0

def to_int(x):
    return int(round(to_float(x)))

def value_band(v):
    v=float(v)
    if v < 10000: return 'Até R$ 10 mil'
    if v < 50000: return 'R$ 10 mil a R$ 50 mil'
    if v < 100000: return 'R$ 50 mil a R$ 100 mil'
    if v < 500000: return 'R$ 100 mil a R$ 500 mil'
    if v < 1000000: return 'R$ 500 mil a R$ 1 milhão'
    return 'Acima de R$ 1 milhão'

def rate_band(v):
    if is_missing(v): return 'Não informado'
    v = to_float(v)
    if v <= 6: return 'Até 6% a.a.'
    if v <= 10: return '6% a 10% a.a.'
    if v <= 14: return '10% a 14% a.a.'
    return 'Acima de 14% a.a.'

def small_count_band(v):
    v=float(v)
    if v <= 0: return '0'
    if v <= 1: return '1'
    if v <= 5: return '2 a 5'
    if v <= 10: return '6 a 10'
    if v <= 50: return '11 a 50'
    return 'Acima de 50'

def month_str(x):
    if is_missing(x): return 'Não informado'
    if hasattr(x, 'strftime'):
        return x.strftime('%Y-%m')
    s = str(x).strip()
    m = re.match(r'^(\d{1,2})/(\d{4})$', s)
    if m: return f"{m.group(2)}-{int(m.group(1)):02d}"
    m = re.match(r'^(\d{4})-(\d{1,2})', s)
    if m: return f"{m.group(1)}-{int(m.group(2)):02d}"
    return 'Não informado'

def add3(dic, key, valor, beneficiarios, contratos):
    arr = dic[key]
    arr[0] += valor; arr[1] += beneficiarios; arr[2] += contratos

def read_headers(ws):
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else '' for h in next(rows)]
    return headers, rows

# Tipology
print('Lendo tipologia...', flush=True)
wb = load_workbook(tipo_path, read_only=True, data_only=True)
ws = wb['Base de dados']
headers, rows = read_headers(ws)
idx = {h:i for i,h in enumerate(headers)}
tipo_map={}; mun_meta={}
for row in rows:
    cod = clean_code(row[idx['Código de município']])
    if not cod: continue
    nome = title_text(row[idx['Nome de município']])
    uf = norm_text(row[idx['UF']]).upper()
    tip = norm_text(row[idx['Tipologia Amazônia Azul']])
    macro = macro_map.get(uf, 'Não informado')
    tipo_map[cod] = tip
    mun_meta[cod] = {'cod_mun':cod,'nome_mun':nome,'uf':uf,'macrorregiao_geografica':macro,'tipologia_pndr':'Não informado','elegivel_amazonia_azul':'Sim','tipologia_territorial_amazonia_azul':tip}
wb.close()
eligible_codes=set(tipo_map.keys())
print('Municípios elegíveis', len(eligible_codes), flush=True)

cube_dims = [
    'ano_mes','fundo_origem','macrorregiao_geografica','uf','cod_mun',
    'elegivel_amazonia_azul','tipologia_territorial_amazonia_azul',
    'atividade_vinculada_amazonia_azul','setor','programa','linha_financiamento','atividade','cnae',
    'porte','finalidade','natureza_contratante','sexo_padronizado','instituicao',
    'faixa_valor_contratado','faixa_taxa_juros','faixa_contratos','faixa_beneficiarios'
]

cube=defaultdict(lambda:[0.0,0,0])
series=defaultdict(lambda:[0.0,0,0])
terr=defaultdict(lambda:[0.0,0,0])
dim_keys=['setor','programa','atividade','cnae','porte','finalidade','uf','macrorregiao_geografica','tipologia_territorial_amazonia_azul','instituicao','natureza_contratante']
dim_agg={d:defaultdict(lambda:[0.0,0,0]) for d in dim_keys}
# benchmarking structure total, azul, fem arrays flattened
bench=defaultdict(lambda:{'total':[0.0,0,0],'azul':[0.0,0,0],'mulheres':[0.0,0,0]})
months=set(); row_count=0

print('Lendo e agregando FNE...', flush=True)
t0=time.time()
wb=load_workbook(fne_path, read_only=True, data_only=True)
ws=wb['Dados_Anexo-I']
headers, rows = read_headers(ws)
idx={h:i for i,h in enumerate(headers)}
for row in rows:
    row_count += 1
    uf = norm_text(row[idx['UF']]).upper()
    cod = clean_code(row[idx['CÓDIGO DE MUNICÍPIO']])
    nome = title_text(row[idx['NOME DO MUNICÍPIO']])
    macro = macro_map.get(uf, 'Não informado')
    tip_pndr = norm_text(row[idx['TIPOLOGIA MUNICÍPIO']])
    ativ_vinc = yes_no(row[idx['TIPOLOGIA AMAZÔNIA AZUL']])
    eleg = 'Sim' if cod in eligible_codes else 'Não'
    tip_terr = tipo_map.get(cod, 'Não elegível')
    ano_mes = month_str(row[idx['DATA DA CONTRATAÇÃO']]); months.add(ano_mes)
    valor = to_float(row[idx['VALOR CONTRATADO']])
    beneficiarios = to_int(row[idx['QTDE DE BENEFICIÁRIOS']])
    contratos = to_int(row[idx['QUANTIDADE DE CONTRATOS']])
    taxa = row[idx['TAXA DE JUROS']]
    data = {
        'ano_mes':ano_mes,
        'fundo_origem':'FNE',
        'macrorregiao_geografica':macro,
        'uf':uf,
        'cod_mun':cod,
        'elegivel_amazonia_azul':eleg,
        'tipologia_territorial_amazonia_azul':tip_terr,
        'atividade_vinculada_amazonia_azul':ativ_vinc,
        'setor':norm_text(row[idx['SETOR']]),
        'programa':norm_text(row[idx['PROGRAMA']]),
        'linha_financiamento':norm_text(row[idx['LINHA DE FINANCIAMENTO']]),
        'atividade':norm_text(row[idx['ATIVIDADE']]),
        'cnae':norm_text(row[idx['CNAE']]),
        'porte':norm_text(row[idx['PORTE']]),
        'finalidade':norm_text(row[idx['FINALIDADE DA OPERAÇÃO']]),
        'natureza_contratante':norm_text(row[idx['PESSOA FISICA JURÍDICA']]),
        'sexo_padronizado':sex_std(row[idx['SEXO']]),
        'instituicao':norm_text(row[idx['INSTITUIÇÃO OPERADORA']]),
        'faixa_valor_contratado':value_band(valor),
        'faixa_taxa_juros':rate_band(taxa),
        'faixa_contratos':small_count_band(contratos),
        'faixa_beneficiarios':small_count_band(beneficiarios)
    }
    key=tuple(data[d] for d in cube_dims); add3(cube,key,valor,beneficiarios,contratos)
    add3(series,(ano_mes,eleg,ativ_vinc,data['sexo_padronizado'],tip_terr,macro,uf),valor,beneficiarios,contratos)
    add3(terr,(ano_mes,cod,uf,nome,macro,eleg,tip_terr,ativ_vinc,data['sexo_padronizado']),valor,beneficiarios,contratos)
    for d in dim_keys:
        add3(dim_agg[d],(data[d],ativ_vinc,data['sexo_padronizado']),valor,beneficiarios,contratos)
    # bench
    bench_keys=[('Brasil','Brasil'),('Macrorregião',macro),('UF',uf),('Tipologia territorial Amazônia Azul',tip_terr),('Município',cod)]
    if eleg=='Sim': bench_keys.append(('Total dos municípios elegíveis','Municípios elegíveis'))
    for bk in bench_keys:
        b=bench[bk]
        b['total'][0]+=valor; b['total'][1]+=beneficiarios; b['total'][2]+=contratos
        if ativ_vinc=='Sim':
            b['azul'][0]+=valor; b['azul'][1]+=beneficiarios; b['azul'][2]+=contratos
        if data['sexo_padronizado']=='Mulheres':
            b['mulheres'][0]+=valor; b['mulheres'][1]+=beneficiarios; b['mulheres'][2]+=contratos
    if cod and cod not in mun_meta:
        mun_meta[cod]={'cod_mun':cod,'nome_mun':nome,'uf':uf,'macrorregiao_geografica':macro,'tipologia_pndr':tip_pndr,'elegivel_amazonia_azul':eleg,'tipologia_territorial_amazonia_azul':tip_terr}
    if row_count % 50000 == 0:
        print('  processadas', row_count, 'linhas em', round(time.time()-t0,1), 's', 'cube', len(cube), flush=True)
wb.close()
print('Linhas processadas', row_count, 'tempo', round(time.time()-t0,1),'s', 'cubo', len(cube), flush=True)

# encode cube columnar
dicts={d:{} for d in cube_dims}; dict_values={d:[] for d in cube_dims}; columns={d:[] for d in cube_dims}; metrics={'valor':[],'beneficiarios':[],'contratos':[]}
def code_for(d, val):
    m=dicts[d]
    if val not in m:
        m[val]=len(m); dict_values[d].append(val)
    return m[val]
for key, vals in cube.items():
    for i,d in enumerate(cube_dims):
        columns[d].append(code_for(d,key[i]))
    metrics['valor'].append(round(vals[0],2)); metrics['beneficiarios'].append(int(vals[1])); metrics['contratos'].append(int(vals[2]))
# sort dictionary values not sorted by occurrence, ok
metadata={'title':'Dashboard Monitoramento Amazônia Azul V2','generated_at':datetime.now().isoformat(timespec='seconds'),'source_files':['FNE_AnexoI_Contratações_032026.xlsx','Tipologia Amazônia Azul (v5).xlsx'],'record_type':'aggregated_cube_dictionary_encoded_columnar','raw_records_not_published':True,'rows_original':row_count,'rows_aggregated':len(cube),'months':sorted(months),'metrics':['valor','beneficiarios','contratos'],'dim_columns':cube_dims,'metric_columns':['valor','beneficiarios','contratos']}
compact={'meta':metadata,'dimensions':cube_dims,'dicts':dict_values,'columns':columns,'metrics':metrics}
(base_dir/'data/processed/operacoes_agregadas_publicas.json').write_bytes(orjson.dumps(compact))

# aux files
def rows_from_agg(dic, headers):
    out=[]
    for key, vals in dic.items():
        rec={h:key[i] for i,h in enumerate(headers)}
        rec.update({'valor':round(vals[0],2),'beneficiarios':int(vals[1]),'contratos':int(vals[2])})
        out.append(rec)
    return out
series_headers=['ano_mes','elegivel_amazonia_azul','atividade_vinculada_amazonia_azul','sexo_padronizado','tipologia_territorial_amazonia_azul','macrorregiao_geografica','uf']
(base_dir/'data/processed/series_temporais.json').write_bytes(orjson.dumps(rows_from_agg(series,series_headers)))
terr_headers=['ano_mes','cod_mun','uf','nome_mun','macrorregiao_geografica','elegivel_amazonia_azul','tipologia_territorial_amazonia_azul','atividade_vinculada_amazonia_azul','sexo_padronizado']
(base_dir/'data/processed/agregados_territoriais.json').write_bytes(orjson.dumps(rows_from_agg(terr,terr_headers)))
dim_out={}
for d,dic in dim_agg.items():
    dim_out[d]=rows_from_agg(dic,[d,'atividade_vinculada_amazonia_azul','sexo_padronizado'])
(base_dir/'data/processed/agregados_dimensoes.json').write_bytes(orjson.dumps(dim_out))
# benchmark records
mun_label={cod:f"{m['nome_mun']} ({m['uf']})" for cod,m in mun_meta.items()}
bench_records=[]
for (typ,label), b in bench.items():
    territory=mun_label.get(label,label) if typ=='Município' else label
    total=b['total']; azul=b['azul']; fem=b['mulheres']
    tv,tb,tc=total; av,ab,ac=azul; fv,fb,fc=fem
    bench_records.append({'tipo_territorio':typ,'territorio':territory,'valor_total':round(tv,2),'beneficiarios_total':int(tb),'contratos_total':int(tc),'valor_azul':round(av,2),'beneficiarios_azul':int(ab),'contratos_azul':int(ac),'participacao_valor_azul':round(av/tv,6) if tv else None,'participacao_beneficiarios_azul':round(ab/tb,6) if tb else None,'participacao_contratos_azul':round(ac/tc,6) if tc else None,'participacao_valor_mulheres':round(fv/tv,6) if tv else None,'participacao_beneficiarios_mulheres':round(fb/tb,6) if tb else None,'participacao_contratos_mulheres':round(fc/tc,6) if tc else None})
(base_dir/'data/processed/benchmarking.json').write_bytes(orjson.dumps(bench_records))
# metadata
municipios=sorted(mun_meta.values(), key=lambda x:(x['uf'],x['nome_mun']))
options={'months':sorted(months),'municipios':municipios,'tipologias_territoriais':sorted(set(tipo_map.values())),'macro_map':macro_map,'source_summary':{'linhas_base_financiamento':row_count,'municipios_tipologia':len(eligible_codes),'municipios_base_financiamento':len(set([m['cod_mun'] for m in mun_meta.values()])), 'mes_inicial':min(months), 'mes_final':max(months)}}
(base_dir/'data/processed/metadata.json').write_bytes(orjson.dumps(options))
manifest={'dashboard':'Dashboard Monitoramento Amazônia Azul V2','version':'2.0','generated_at':metadata['generated_at'],'files':{'operacoes_agregadas_publicas':'data/processed/operacoes_agregadas_publicas.json','series_temporais':'data/processed/series_temporais.json','agregados_territoriais':'data/processed/agregados_territoriais.json','agregados_dimensoes':'data/processed/agregados_dimensoes.json','benchmarking':'data/processed/benchmarking.json','metadata':'data/processed/metadata.json'},'no_raw_data':True,'raw_excluded_patterns':['data/raw/','*.xlsx','*.xls','*.csv']}
(base_dir/'data/manifest.json').write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding='utf-8')
geo_manifest={'package':'dashboard_monitoramento_amazul_v2','maps':{'local_topojson_supported':['data/geo/municipios_ibge_topo.json','data/geo/ufs_ibge_topo.json'],'remote_fallback':{'municipios':'https://servicodados.ibge.gov.br/api/v3/malhas/paises/BR?intrarregiao=municipio&qualidade=minima&formato=application/json','ufs':'https://servicodados.ibge.gov.br/api/v3/malhas/paises/BR?intrarregiao=UF&qualidade=minima&formato=application/json'}},'notes':'O dashboard tenta carregar TopoJSON local. Se ausente, usa as malhas simplificadas oficiais do IBGE via API. Para funcionamento 100% offline, copie os arquivos municipios_ibge_topo.json e ufs_ibge_topo.json para esta pasta.'}
(base_dir/'data/geo/manifest_malhas.json').write_text(json.dumps(geo_manifest, ensure_ascii=False, indent=2), encoding='utf-8')
print('Tamanhos JSON:', flush=True)
for p in sorted((base_dir/'data/processed').glob('*.json')):
    print(p.name, round(p.stat().st_size/1024/1024,2),'MB', flush=True)
print('OK', base_dir, flush=True)
